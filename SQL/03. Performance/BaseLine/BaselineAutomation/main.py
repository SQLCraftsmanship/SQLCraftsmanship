import os
import sys
import pyodbc
import pandas as pd
import openpyxl
import logging
import json
import re
from colorama import init, Fore, Style
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

init(autoreset=True)

# ------------------- Python execution -------------------------
# Windows Authentication (Auto-detect)
# python main.py
# Windows Authentication (Specify SQL Server)
# python main.py -parSQLInstanceName "SQLSERVER01"

# SQL Authentication
#python main.py -parSQLInstanceName "SQLSERVER01" -parUserName "sa" -useSQLAuth -parPassword "MyPassword123"

# ------------------- Python General -------------------------
# Setup Script for Packaging as Executable
# pip install pyinstaller

# To create a .exe
# python -m PyInstaller --onefile --noconsole main.py


# ------------------- General Configurations -------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SCRIPTS_DIR = os.path.join(BASE_DIR, 'Script', 'sps')
EXECUTE_SCRIPT_PATH = os.path.join(BASE_DIR, 'Script', 'execute', 'run.sql')
OUTPUT_DIR = os.path.join(BASE_DIR, 'Output')
LOG_DIR = os.path.join(BASE_DIR, 'Log')
FORMAT_CONFIG_PATH = os.path.join(BASE_DIR, 'FormatConfig.json')

# print(f"****** {SCRIPTS_DIR}  ******")
# print(f"****** {EXECUTE_SCRIPT_PATH} ******")

if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)
if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)

log_filename = os.path.join(LOG_DIR, 'execution_log.txt')
logging.basicConfig(filename=log_filename, level=logging.INFO, format='%(asctime)s %(message)s')

# ------------------- Utility Functions -------------------
def get_sql_server_instance():
    return os.environ.get('COMPUTERNAME', 'localhost')

def get_windows_user():
    return os.environ.get('USERNAME', 'sa')

def print_success(message):
    print(Fore.GREEN + message)

def print_error(message):
    print(Fore.RED + message)

def connect_to_sql_server(instance_name=None, user_name=None, use_sql_auth=False, password=None):
    try:
        if instance_name is None:
            instance_name = get_sql_server_instance()
        if user_name is None:
            user_name = get_windows_user()
        if use_sql_auth:
            conn_str = f'DRIVER={{SQL Server}};SERVER={instance_name};UID={user_name};PWD={password}'
        else:
            conn_str = f'DRIVER={{SQL Server}};SERVER={instance_name};Trusted_Connection=yes;'
            
        # conn_str = f'DRIVER={{SQL Server}};SERVER={instance_name};Trusted_Connection=yes;'
        conn = pyodbc.connect(conn_str)
        print_success(f"***** Successfully connected to SQL Server Instance: {instance_name} *****")
        return conn
    except Exception as e:
        print_error(f"***** Error connecting to SQL Server Instance: {e} *****")
        sys.exit(1)

# ------------------- Script Execution ---------------------
def execute_scripts_in_folder(cursor, folder_path):
    scripts = sorted([f for f in os.listdir(folder_path) if f.endswith('.sql')])
    for script in scripts:
        script_path = os.path.join(folder_path, script)
        print_success(f"***** Executing script: {script} *****")
        try:
            with open(script_path, 'r') as file:
                sql_commands = file.read().split('GO')
                for command in sql_commands:
                    if command.strip():
                        cursor.execute(command)
            print_success(f"***** Script {script} executed successfully. *****")
            logging.info(f"The script {script} was executed successfully.")
        except Exception as e:
            print_error(f"***** Script {script} failed: {e} *****")
            logging.error(f"The script {script} failed. Error: {e}")
            sys.exit(1)

def execute_run_sql(cursor, conn, run_sql_path):
    try:
        # ... BaseLine\BaselineAutomation\Script\execute\run.sql
        # print(f"****** The run_sql_path is: {run_sql_path}  ******")
        
        with open(run_sql_path, 'r') as file:
            sql_commands = file.read().split('GO')
            
            # The sql_commands is: ['EXEC [master].[dbo].[sp_GetLastReboot]\n', '\n\nEXEC [master].[dbo].[sp_OSInfo]\n', '\n\nEXEC [master].[dbo].[sp_SQLServerProperty]\n', '\n\nEXEC [master].[dbo].[sp_SQLServiceInfo]\n', '\n\nEXEC [master].[dbo].[sp_SQLVersionInfo]\n', ''] 
            # print(f"****** The sql_commands is: {sql_commands}  ******")

        for idx, command in enumerate(sql_commands):
            if command.strip():
                try:
                    print_success(f"***** Executing stored procedure block {idx + 1}... *****")
                    cursor.execute(command)

                    # The command is: EXEC [master].[dbo].[sp_GetLastReboot]
                    # Extract the stored procedure name using regex
                    spName = command.split('[master].')[1].strip() + '.sql'
                    print(spName)
                        
                    # save_results_to_excel(cursor, f'SP_{idx + 1}')
                    save_results_to_excel(cursor, spName)
                    
                    conn.commit()
                    print_success(f"***** Block {idx + 1} executed successfully. *****")
                    logging.info(f"Block {idx + 1} executed successfully.")
                except Exception as e:
                    print_error(f"***** Error executing block {idx + 1}: {e} *****")
                    logging.error(f"Error executing block {idx + 1}: {e}")
                    continue
    except Exception as e:
        print_error(f"***** Error reading run.sql: {e} *****")
        sys.exit(1)

# ------------------- JSON Configuration -------------------
def load_json_config():
    with open(FORMAT_CONFIG_PATH, 'r') as file:
        return json.load(file)

# ------------------- Excel Export -------------------------
def save_results_to_excel(cursor, script_name):
    config = load_json_config()
    if script_name not in config:
        print_error(f"***** No configuration found for {script_name} in FormatConfig.json. *****")
        return

    tab_config = config[script_name]
    tab_name = tab_config['tab_name']
    header_color = tab_config['header_color']
    tab_color = tab_config['tab_color']
    headers = tab_config['column_headers']

    output_path = os.path.join(OUTPUT_DIR, 'BaselineResults.xlsx')

    if os.path.exists(output_path):
        workbook = openpyxl.load_workbook(output_path)
    else:
        workbook = openpyxl.Workbook()
        del workbook[workbook.sheetnames[0]]

    sheet = workbook.create_sheet(title=tab_name)
    sheet.sheet_properties.tabColor = tab_color

    result_sets = []
    while True:
        columns = [column[0] for column in cursor.description]
        rows = cursor.fetchall()
        result_sets.append((columns, rows))
        if not cursor.nextset():
            break

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    row_index = 1
    for columns, rows in result_sets:
        if rows:
            for col_index, header in enumerate(headers, 1):
                cell = sheet.cell(row=row_index, column=col_index, value=header)
                # cell.fill = openpyxl.styles.PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            row_index += 1
            
            for row in rows:
                for col_index, value in enumerate(row, 1):
                    # sheet.cell(row=row_index, column=col_index, value=value)
                    cell = sheet.cell(row=row_index, column=col_index, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = thin_border
                row_index += 1

            row_index += 1

    workbook.save(output_path)
    print_success(f"****** Results for {script_name} exported successfully to Excel. ******")
    logging.info(f"Results for {script_name} exported successfully to Excel.")

# ------------------- Main Execution -------------------
def main():
    sql_instance = None
    user_name = None
    use_sql_auth = False
    password = None
    
    if '-parSQLInstanceName' in sys.argv:
        sql_instance = sys.argv[sys.argv.index('-parSQLInstanceName') + 1]

    if '-parUserName' in sys.argv:
        user_name = sys.argv[sys.argv.index('-parUserName') + 1]

    if '-useSQLAuth' in sys.argv:
        use_sql_auth = True

    if '-parPassword' in sys.argv:
        password = sys.argv[sys.argv.index('-parPassword') + 1]

    conn = connect_to_sql_server(sql_instance, user_name, use_sql_auth, password)
    cursor = conn.cursor()

    execute_scripts_in_folder(cursor, SCRIPTS_DIR)
    print_success("***** All stored procedures created successfully *****")

    execute_run_sql(cursor, conn, EXECUTE_SCRIPT_PATH)
    print_success("***** All stored procedures executed successfully *****")

    conn.close()

if __name__ == '__main__':
    main()
