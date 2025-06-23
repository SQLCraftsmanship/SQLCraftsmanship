import os
import sys
import pyodbc
import pandas as pd
import json
import logging
from datetime import datetime
from colorama import init, Fore, Style
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Initialize colorama
init(autoreset=True)

# ===========================
# Configuration
# ===========================
SCRIPT_FOLDER = 'Script'
OUTPUT_FOLDER = 'Output'
LOG_FOLDER = 'Log'
CONFIG_FILE = 'FormatConfig.json'

# ===========================
# Database Connection
# ===========================
def get_db_connection():
    # You can replace this with reading from a config file or parsing CLI args
    conn_str = (
        "DRIVER={ODBC Driver 17 for SQL Server};"
        "SERVER=.\SQL2022DEV;"
        "DATABASE=master;"
        "Trusted_Connection=yes;"
    )
    return pyodbc.connect(conn_str)

# ===========================
# Logging Setup
# ===========================
def setup_logger():
    if not os.path.exists(LOG_FOLDER):
        os.makedirs(LOG_FOLDER)
    log_filename = f"Baseline_Log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    log_path = os.path.join(LOG_FOLDER, log_filename)
    logging.basicConfig(filename=log_path, level=logging.INFO, format='%(asctime)s - %(message)s')
    return log_path

# ===========================
# Load Format Config
# ===========================
def load_format_config():
    with open(CONFIG_FILE, 'r') as f:
        return json.load(f)

# ===========================
# Execute SQL Script
# ===========================
def execute_sql_script(cursor, script_path):
    with open(script_path, 'r', encoding='utf-8') as file:
        sql = file.read()
    cursor.execute(sql)
    try:
        columns = [column[0] for column in cursor.description]
        rows = cursor.fetchall()
        df = pd.DataFrame.from_records(rows, columns=columns)
        return df
    except Exception:
        # Script executed but no result set (for example DML statements)
        return pd.DataFrame()

# ===========================
# Apply Header Formatting
# ===========================
def apply_header_formatting(ws, header_color):
    fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
    for cell in ws[1]:
        cell.fill = fill

# ===========================
# Main Process
# ===========================
def main():
    # Setup
    if not os.path.exists(OUTPUT_FOLDER):
        os.makedirs(OUTPUT_FOLDER)
    log_path = setup_logger()
    config = load_format_config()

    excel_filename = f"Baseline_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    excel_path = os.path.join(OUTPUT_FOLDER, excel_filename)

    conn = get_db_connection()
    cursor = conn.cursor()

    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    script_files = [f for f in os.listdir(SCRIPT_FOLDER) if f.endswith('.sql')]

    for script_file in script_files:
        script_path = os.path.join(SCRIPT_FOLDER, script_file)
        display_name = f"[{script_file}]"

        print(Fore.GREEN + f"✅ Start: {display_name} is being run.")
        try:
            df = execute_sql_script(cursor, script_path)

            tab_name = config.get(script_file, {}).get('tab_name', script_file.replace('.sql', ''))
            headers = config.get(script_file, {}).get('column_headers', df.columns.tolist())
            header_color = config.get(script_file, {}).get('header_color', 'FFFFFF')

            ws = wb.create_sheet(title=tab_name)

            if df.empty:
                ws.append(['No data returned'])
            else:
                # Apply custom headers from JSON
                ws.append(headers)
                for row in df.itertuples(index=False, name=None):
                    ws.append(row)

                apply_header_formatting(ws, header_color)

            logging.info(f"The script {script_file} was executed successfully.")
            print(Fore.GREEN + f"✅ Success: {display_name} finished successfully.")

        except Exception as e:
            error_message = f"The script {script_file} failed. Error: {str(e)}"
            logging.error(error_message)
            print(Fore.RED + f"❌ Failure: {display_name} finished with error. Please see the Log.")
            continue  # Move to the next script

    wb.save(excel_path)
    print(Fore.CYAN + f"\n🎉 Baseline Excel file generated: {excel_path}")
    print(Fore.CYAN + f"📄 Log file saved at: {log_path}")
    conn.close()

if __name__ == "__main__":
    main()
